"""Excel：各 scope 模板下載與批次上傳（all-or-nothing）。"""
import io
import json as jsonlib
from typing import Any

import yaml as yamllib
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ...database import CompanyMaster, ProviderRecord, ScopeDef, SystemParam, UserPref, get_db

from .core import _scope_schema, _validate_row
from .records import upsert_master, upsert_param, upsert_provider

router = APIRouter()


@router.get("/{scope}/template")
def download_template(scope: str, db: Session = Depends(get_db)):
    schema = _scope_schema(scope, db)
    wb = Workbook()
    ws = wb.active
    ws.title = schema["label"][:31]
    keys = [f["key"] for f in schema["fields"]]
    ws.append(keys)
    ws.append([("必填" if f.get("required") else "選填") + (f" — {f.get('hint','')}" if f.get("hint") else "")
               for f in schema["fields"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{scope}_template.xlsx"'})


@router.post("/{scope}/upload")
async def batch_upload(scope: str, file: UploadFile, db: Session = Depends(get_db)):
    """Excel 批次 upsert。第一列為欄位名；回傳逐列驗證報告，全部通過才寫入（交易一致性）。"""
    schema = _scope_schema(scope, db)
    try:
        wb = load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(422, "無法解析 Excel 檔案，請使用模板格式（.xlsx）")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(422, "空白檔案")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    known = {f["key"] for f in schema["fields"]}
    unknown = [h for h in header if h and h not in known]

    report: list[dict] = []
    valid_rows: list[dict] = []
    for idx, raw in enumerate(rows[1:], start=2):
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        data = {header[i]: raw[i] for i in range(min(len(header), len(raw))) if header[i]}
        # 跳過模板第二列的說明列
        if idx == 2 and any(isinstance(v, str) and ("必填" in v or "選填" in v) for v in data.values()):
            continue
        data = {k: (str(v).strip() if v is not None else None) for k, v in data.items()}
        errs = _validate_row(data, schema["fields"], db,
                             check_fk=(db.get(ScopeDef, scope) is not None and scope not in ("company_master", "system_params")))
        report.append({"row": idx, "key": data.get(schema.get("key_field", "fab_code")), "errors": errs})
        if not errs:
            valid_rows.append(data)

    failed = [r for r in report if r["errors"]]
    if failed:
        return {"status": "rejected", "message": f"{len(failed)} 列驗證失敗，全部未寫入",
                "unknown_columns": unknown, "report": report}

    key_field = schema.get("key_field", "fab_code")
    for data in valid_rows:
        if scope == "company_master":
            upsert_master(data[key_field], data, db)
        elif scope == "system_params":
            upsert_param(data[key_field], data, db)
        else:
            upsert_provider(scope, data["fab_code"], data, db)
    return {"status": "ok", "upserted": len(valid_rows), "unknown_columns": unknown, "report": report}


