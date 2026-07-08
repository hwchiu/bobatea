"""Settings — Company Config（Company Master + Data Providers）與 System Parameters。

設計原則：
1. Schema-driven：PROVIDER_SCHEMAS 定義各 data provider 需要維護的欄位，
   前端表格欄位、Excel 模板、上傳驗證全部由此驅動；新增 provider 僅需加一筆設定。
2. Company Master 為主檔：所有 provider mapping 的 fab_code 都必須先存在於主檔（FK 驗證）。
3. 儲存層走 SQLAlchemy（目前 SQLite，切 PostgreSQL / MariaDB 只需改 DATABASE_URL）。
"""
import io
import json as jsonlib

import yaml as yamllib
from datetime import datetime
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..database import CompanyMaster, ProviderRecord, ScopeDef, SystemParam, UserPref, get_db

router = APIRouter()

# ── Schema 註冊表 ─────────────────────────────────────────────
# field: {key, label, required, type(text|number|select), options?, hint?}

DEFAULT_COMPANY_MASTER = {
    "id": "company_master",
    "label": "Company Master",
    "key_field": "fab_code",
    "fields": [
        {"key": "fab_code", "label": "Fab Code", "required": True, "type": "text",
         "hint": "公司主檔唯一代碼（primary key），所有 provider mapping 以此關聯"},
        {"key": "company_name", "label": "Company Name", "required": True, "type": "text"},
        {"key": "company_short_name", "label": "Short Name", "required": True, "type": "text"},
        {"key": "country", "label": "Country", "required": False, "type": "text"},
        {"key": "region", "label": "Region", "required": False, "type": "select",
         "options": ["APAC", "EMEA", "AMER"]},
        {"key": "status", "label": "Status", "required": False, "type": "select",
         "options": ["active", "inactive"], "default": "active"},
        {"key": "remark", "label": "Remark", "required": False, "type": "text"},
    ],
}

DEFAULT_PROVIDERS: dict[str, dict[str, Any]] = {
    "bloomberg": {
        "id": "bloomberg", "label": "Bloomberg",
        "fields": [
            {"key": "fab_code", "label": "Fab Code", "required": True, "type": "fk"},
            {"key": "bbg_id", "label": "BBG ID (FIGI)", "required": True, "type": "text",
             "hint": "Bloomberg Global Identifier，如 BBG000B9XRY4"},
            {"key": "ticker", "label": "Ticker", "required": True, "type": "text"},
            {"key": "exchange_code", "label": "Exchange Code", "required": False, "type": "text",
             "hint": "如 TT / US / JP"},
            {"key": "currency", "label": "Currency", "required": False, "type": "text"},
            {"key": "price_offset_days", "label": "Offset (days)", "required": False, "type": "number",
             "hint": "取價日偏移，處理時差 / 結算延遲"},
            {"key": "status", "label": "Status", "required": False, "type": "select",
             "options": ["active", "inactive"], "default": "active"},
        ],
    },
    "factset": {
        "id": "factset", "label": "FactSet",
        "fields": [
            {"key": "fab_code", "label": "Fab Code", "required": True, "type": "fk"},
            {"key": "factset_entity_id", "label": "FactSet ID", "required": True, "type": "text",
             "hint": "如 0FPWZZ-E"},
            {"key": "fsym_id", "label": "FSYM ID", "required": False, "type": "text",
             "hint": "FactSet permanent security identifier"},
            {"key": "ticker_region", "label": "Ticker-Region", "required": False, "type": "text",
             "hint": "如 2330-TW"},
            {"key": "status", "label": "Status", "required": False, "type": "select",
             "options": ["active", "inactive"], "default": "active"},
        ],
    },
    "contify": {
        "id": "contify", "label": "Contify",
        "fields": [
            {"key": "fab_code", "label": "Fab Code", "required": True, "type": "fk"},
            {"key": "company_name", "label": "Company Name", "required": True, "type": "text",
             "hint": "新聞追蹤顯示名稱，供關鍵字 / watchlist 建檔使用"},
            {"key": "contify_company_id", "label": "Contify Company ID", "required": True, "type": "text"},
            {"key": "watchlist_id", "label": "Watchlist ID", "required": False, "type": "text"},
            {"key": "topics", "label": "Topics", "required": False, "type": "text",
             "hint": "逗號分隔的追蹤主題，如 M&A,Capacity Expansion"},
            {"key": "status", "label": "Status", "required": False, "type": "select",
             "options": ["active", "inactive"], "default": "active"},
        ],
    },
    "snp": {
        "id": "snp", "label": "S&P",
        "fields": [
            {"key": "fab_code", "label": "Fab Code", "required": True, "type": "fk"},
            {"key": "spciq_id", "label": "S&P Capital IQ ID", "required": True, "type": "text",
             "hint": "如 IQ24937"},
            {"key": "gvkey", "label": "GVKEY", "required": False, "type": "text",
             "hint": "Compustat 公司永久代碼"},
            {"key": "snp_ticker", "label": "S&P Ticker", "required": False, "type": "text"},
            {"key": "status", "label": "Status", "required": False, "type": "select",
             "options": ["active", "inactive"], "default": "active"},
        ],
    },
    "dnb": {
        "id": "dnb", "label": "D&B",
        "fields": [
            {"key": "fab_code", "label": "Fab Code", "required": True, "type": "fk"},
            {"key": "duns_number", "label": "DUNS Number", "required": True, "type": "text",
             "hint": "9 碼 D-U-N-S 編號"},
            {"key": "global_ultimate_duns", "label": "Global Ultimate DUNS", "required": False, "type": "text",
             "hint": "全球最終母公司 DUNS"},
            {"key": "tradestyle_name", "label": "Tradestyle Name", "required": False, "type": "text"},
            {"key": "status", "label": "Status", "required": False, "type": "select",
             "options": ["active", "inactive"], "default": "active"},
        ],
    },
}

DEFAULT_SYSTEM_PARAMS = {
    "id": "system_params",
    "label": "System Parameter",
    "key_field": "param_key",
    "fields": [
        {"key": "param_key", "label": "Key", "required": True, "type": "text"},
        {"key": "param_value", "label": "Value", "required": True, "type": "text"},
        {"key": "category", "label": "Category", "required": False, "type": "select",
         "options": ["scheduler", "ai_kernel", "data_platform", "general"], "default": "general"},
        {"key": "value_type", "label": "Type", "required": False, "type": "select",
         "options": ["string", "number", "boolean", "json"], "default": "string"},
        {"key": "description", "label": "Description", "required": False, "type": "text"},
    ],
}


@router.get("/schema")
def get_schema(db: Session = Depends(get_db)):
    """前端動態渲染表格 / 表單 / Excel 欄位的單一來源（DB-backed，Admin 可調整）。"""
    _ensure_seed(db)
    providers = {r.scope_id: _row_to_schema(r) for r in
                 db.query(ScopeDef).filter(ScopeDef.group == "company_config",
                                           ScopeDef.scope_id != "company_master")
                   .order_by(ScopeDef.scope_id).all()}
    return {
        "company_master": _scope("company_master", db),
        "providers": providers,
        "system_params": _scope("system_params", db),
    }


# ── Company Master CRUD ──────────────────────────────────────

@router.get("/company-master")
def list_master(db: Session = Depends(get_db)):
    rows = db.query(CompanyMaster).order_by(CompanyMaster.fab_code).all()
    return [_master_out(r) for r in rows]


@router.put("/company-master/{fab_code}")
def upsert_master(fab_code: str, body: dict[str, Any], db: Session = Depends(get_db)):
    errs = _validate_row(body | {"fab_code": fab_code}, _scope("company_master", db)["fields"], db, check_fk=False)
    if errs:
        raise HTTPException(422, errs)
    row = db.get(CompanyMaster, fab_code)
    if not row:
        row = CompanyMaster(fab_code=fab_code)
        db.add(row)
    for f in _scope("company_master", db)["fields"]:
        if f["key"] != "fab_code" and f["key"] in body:
            setattr(row, f["key"], body[f["key"]])
    db.commit()
    return _master_out(row)


@router.delete("/company-master/{fab_code}")
def delete_master(fab_code: str, db: Session = Depends(get_db)):
    used = db.query(ProviderRecord).filter(ProviderRecord.fab_code == fab_code).count()
    if used:
        raise HTTPException(409, f"fab_code {fab_code} 仍被 {used} 筆 provider mapping 引用，請先刪除相關 mapping")
    row = db.get(CompanyMaster, fab_code)
    if not row:
        raise HTTPException(404, "not found")
    db.delete(row)
    db.commit()
    return {"deleted": fab_code}


# ── Provider mapping CRUD ────────────────────────────────────

def _provider_or_404(provider: str, db: Session) -> dict:
    row = db.get(ScopeDef, provider)
    if not row or row.group != "company_config" or provider == "company_master":
        raise HTTPException(404, f"未知的 provider: {provider}")
    return _row_to_schema(row)


@router.get("/providers/{provider}/records")
def list_provider(provider: str, db: Session = Depends(get_db)):
    _provider_or_404(provider, db)
    rows = (db.query(ProviderRecord).filter(ProviderRecord.provider == provider)
            .order_by(ProviderRecord.fab_code).all())
    return [_record_out(r) for r in rows]


@router.put("/providers/{provider}/records/{fab_code}")
def upsert_provider(provider: str, fab_code: str, body: dict[str, Any], db: Session = Depends(get_db)):
    schema = _provider_or_404(provider, db)
    data = body | {"fab_code": fab_code}
    errs = _validate_row(data, schema["fields"], db, check_fk=True)
    if errs:
        raise HTTPException(422, errs)
    row = (db.query(ProviderRecord)
           .filter(ProviderRecord.provider == provider, ProviderRecord.fab_code == fab_code).first())
    if not row:
        row = ProviderRecord(provider=provider, fab_code=fab_code, attrs={})
        db.add(row)
    attrs = dict(row.attrs or {})
    for f in schema["fields"]:
        if f["key"] != "fab_code" and f["key"] in data:
            attrs[f["key"]] = data[f["key"]]
    row.attrs = attrs
    db.commit()
    return _record_out(row)


@router.delete("/providers/{provider}/records/{fab_code}")
def delete_provider(provider: str, fab_code: str, db: Session = Depends(get_db)):
    _provider_or_404(provider, db)
    row = (db.query(ProviderRecord)
           .filter(ProviderRecord.provider == provider, ProviderRecord.fab_code == fab_code).first())
    if not row:
        raise HTTPException(404, "not found")
    db.delete(row)
    db.commit()
    return {"deleted": fab_code}


# ── System parameters CRUD ───────────────────────────────────

@router.get("/system-params")
def list_params(db: Session = Depends(get_db)):
    return [{"param_key": r.param_key, "param_value": r.param_value, "category": r.category,
             "value_type": r.value_type, "description": r.description} for r in
            db.query(SystemParam).order_by(SystemParam.category, SystemParam.param_key).all()]


@router.put("/system-params/{param_key}")
def upsert_param(param_key: str, body: dict[str, Any], db: Session = Depends(get_db)):
    row = db.get(SystemParam, param_key)
    if not row:
        row = SystemParam(param_key=param_key)
        db.add(row)
    for k in ("param_value", "category", "value_type", "description"):
        if k in body:
            setattr(row, k, body[k])
    db.commit()
    return {"param_key": param_key}


@router.delete("/system-params/{param_key}")
def delete_param(param_key: str, db: Session = Depends(get_db)):
    row = db.get(SystemParam, param_key)
    if not row:
        raise HTTPException(404, "not found")
    db.delete(row)
    db.commit()
    return {"deleted": param_key}


# ── Excel：模板下載 + 批次上傳 ────────────────────────────────

def _scope_schema(scope: str, db: Session) -> dict:
    if scope in ("company_master", "system_params"):
        return _scope(scope, db)
    return _provider_or_404(scope, db)


@router.get("/{scope}/template")
def download_template(scope: str, db: Session = Depends(get_db)):
    schema = _scope_schema(scope, db)
    wb = Workbook()
    ws = wb.active
    ws.title = schema["label"][:31]
    keys = [f["key"] for f in schema["fields"]]
    ws.append(keys)
    ws.append([("必填" if f.get("required") else "選填") + (f" — {f.get('hint','')}" if f.get("hint") else "")
               for f in schema["fields"]])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{scope}_template.xlsx"'})


@router.post("/{scope}/upload")
async def batch_upload(scope: str, file: UploadFile, db: Session = Depends(get_db)):
    """Excel 批次 upsert。第一列為欄位名；回傳逐列驗證報告，全部通過才寫入（交易一致性）。"""
    schema = _scope_schema(scope, db)
    try:
        wb = load_workbook(io.BytesIO(await file.read()), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(422, "無法解析 Excel 檔案，請使用模板格式（.xlsx）")
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(422, "空白檔案")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    known = {f["key"] for f in schema["fields"]}
    unknown = [h for h in header if h and h not in known]

    report: list[dict] = []
    valid_rows: list[dict] = []
    for idx, raw in enumerate(rows[1:], start=2):
        if raw is None or all(v is None or str(v).strip() == "" for v in raw):
            continue
        data = {header[i]: raw[i] for i in range(min(len(header), len(raw))) if header[i]}
        # 跳過模板第二列的說明列
        if idx == 2 and any(isinstance(v, str) and ("必填" in v or "選填" in v) for v in data.values()):
            continue
        data = {k: (str(v).strip() if v is not None else None) for k, v in data.items()}
        errs = _validate_row(data, schema["fields"], db,
                             check_fk=(db.get(ScopeDef, scope) is not None and scope not in ("company_master", "system_params")))
        report.append({"row": idx, "key": data.get(schema.get("key_field", "fab_code")), "errors": errs})
        if not errs:
            valid_rows.append(data)

    failed = [r for r in report if r["errors"]]
    if failed:
        return {"status": "rejected", "message": f"{len(failed)} 列驗證失敗，全部未寫入",
                "unknown_columns": unknown, "report": report}

    key_field = schema.get("key_field", "fab_code")
    for data in valid_rows:
        if scope == "company_master":
            upsert_master(data[key_field], data, db)
        elif scope == "system_params":
            upsert_param(data[key_field], data, db)
        else:
            upsert_provider(scope, data["fab_code"], data, db)
    return {"status": "ok", "upserted": len(valid_rows), "unknown_columns": unknown, "report": report}


# ── 共用驗證 / 序列化 ────────────────────────────────────────

def _validate_row(data: dict, fields: list[dict], db: Session, check_fk: bool) -> list[str]:
    errs: list[str] = []
    for f in fields:
        v = data.get(f["key"])
        empty = v is None or str(v).strip() == ""
        if f.get("required") and empty:
            errs.append(f"{f['key']} 為必填")
            continue
        if empty:
            continue
        if f["type"] == "number":
            try:
                float(v)
            except (TypeError, ValueError):
                errs.append(f"{f['key']} 必須為數字")
        if f["type"] == "select" and f.get("options") and str(v) not in f["options"]:
            errs.append(f"{f['key']} 必須為 {f['options']} 之一")
        if f["type"] == "fk" and check_fk:
            if not db.get(CompanyMaster, str(v)):
                errs.append(f"fab_code {v} 不存在於 Company Master，請先維護主檔")
    return errs


def _master_out(r: CompanyMaster) -> dict:
    return {"fab_code": r.fab_code, "company_name": r.company_name,
            "company_short_name": r.company_short_name, "country": r.country,
            "region": r.region, "status": r.status, "remark": r.remark,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None}


def _record_out(r: ProviderRecord) -> dict:
    return {"fab_code": r.fab_code, **(r.attrs or {}),
            "updated_at": r.updated_at.isoformat() if r.updated_at else None}


# ── Scope 動態載入 / seed ─────────────────────────────────────

def _row_to_schema(r: ScopeDef) -> dict:
    return {"id": r.scope_id, "label": r.label, "group": r.group,
            "key_field": r.key_field, "fields": r.fields or [],
            "backend": r.backend or {"type": "rdb"}, "builtin": bool(r.builtin)}


def _scope(scope_id: str, db: Session) -> dict:
    _ensure_seed(db)
    row = db.get(ScopeDef, scope_id)
    if not row:
        raise HTTPException(404, f"scope {scope_id} 不存在")
    return _row_to_schema(row)


def _ensure_seed(db: Session) -> None:
    """首次啟動：預設 schema 寫入 scope_def；之後以 DB 為準（Admin 可調整）。"""
    if db.query(ScopeDef).count() > 0:
        return
    def add(d: dict, group: str, key_field: str):
        db.add(ScopeDef(scope_id=d["id"], group=group, label=d["label"],
                        key_field=key_field, fields=d["fields"],
                        backend={"type": "rdb"}, builtin=1))
    add(DEFAULT_COMPANY_MASTER, "company_config", "fab_code")
    for p in DEFAULT_PROVIDERS.values():
        add(p, "company_config", "fab_code")
    add(DEFAULT_SYSTEM_PARAMS, "system", "param_key")
    db.commit()


# ── Admin：scope / 欄位 / 後端連結管理 ────────────────────────

@router.get("/admin/scopes")
def admin_list_scopes(db: Session = Depends(get_db)):
    _ensure_seed(db)
    return [_row_to_schema(r) for r in db.query(ScopeDef).order_by(ScopeDef.group, ScopeDef.scope_id).all()]


@router.put("/admin/scopes/{scope_id}")
def admin_upsert_scope(scope_id: str, body: dict[str, Any], db: Session = Depends(get_db)):
    """新增或調整 scope：label、fields（欄位定義）、backend（rdb / file:json|yaml）。"""
    _ensure_seed(db)
    if not scope_id.replace("_", "").isalnum():
        raise HTTPException(422, "scope_id 僅允許英數與底線")
    fields = body.get("fields")
    if fields is not None:
        errs = _validate_fields_def(fields)
        if errs:
            raise HTTPException(422, errs)
    backend = body.get("backend")
    if backend is not None:
        if backend.get("type") not in ("rdb", "file"):
            raise HTTPException(422, "backend.type 必須為 rdb 或 file")
        if backend.get("type") == "file" and backend.get("format") not in ("json", "yaml"):
            raise HTTPException(422, "backend.format 必須為 json 或 yaml")

    row = db.get(ScopeDef, scope_id)
    if not row:
        row = ScopeDef(scope_id=scope_id, group=body.get("group", "company_config"),
                       key_field=body.get("key_field", "fab_code"), builtin=0,
                       fields=[], backend={"type": "rdb"})
        db.add(row)
    if row.builtin and body.get("group") and body["group"] != row.group:
        raise HTTPException(409, "內建 scope 不可變更 group")
    for k in ("label", "fields", "backend"):
        if body.get(k) is not None:
            setattr(row, k, body[k])
    if not row.builtin and body.get("key_field"):
        row.key_field = body["key_field"]
    db.commit()
    return _row_to_schema(row)


@router.delete("/admin/scopes/{scope_id}")
def admin_delete_scope(scope_id: str, db: Session = Depends(get_db)):
    row = db.get(ScopeDef, scope_id)
    if not row:
        raise HTTPException(404, "not found")
    if row.builtin:
        raise HTTPException(409, "內建 scope 不可刪除（可調整欄位或後端連結）")
    n = db.query(ProviderRecord).filter(ProviderRecord.provider == scope_id).count()
    if n:
        raise HTTPException(409, f"scope 內尚有 {n} 筆資料，請先清空")
    db.delete(row)
    db.commit()
    return {"deleted": scope_id}


def _validate_fields_def(fields: list) -> list[str]:
    errs: list[str] = []
    if not isinstance(fields, list) or not fields:
        return ["fields 必須為非空陣列"]
    seen: set[str] = set()
    for i, f in enumerate(fields):
        key = (f or {}).get("key", "")
        if not key or not str(key).replace("_", "").isalnum():
            errs.append(f"fields[{i}].key 無效")
        if key in seen:
            errs.append(f"fields[{i}].key 重複: {key}")
        seen.add(key)
        if (f or {}).get("type") not in ("text", "number", "select", "fk"):
            errs.append(f"fields[{i}].type 必須為 text/number/select/fk")
        if f.get("type") == "select" and not f.get("options"):
            errs.append(f"fields[{i}] select 型別需提供 options")
    return errs


# ── 資料輸出：json / yaml ─────────────────────────────────────

def _scope_rows(scope: str, db: Session) -> list[dict]:
    if scope == "company_master":
        return [_master_out(r) for r in db.query(CompanyMaster).order_by(CompanyMaster.fab_code).all()]
    if scope == "system_params":
        return [{"param_key": r.param_key, "param_value": r.param_value, "category": r.category,
                 "value_type": r.value_type, "description": r.description}
                for r in db.query(SystemParam).order_by(SystemParam.param_key).all()]
    return [_record_out(r) for r in db.query(ProviderRecord)
            .filter(ProviderRecord.provider == scope).order_by(ProviderRecord.fab_code).all()]


@router.get("/{scope}/export")
def export_scope(scope: str, format: str = "json", db: Session = Depends(get_db)):
    """scope 資料輸出 json / yaml（後端連結 type=file 的落地格式，也可隨選下載）。"""
    _scope_schema(scope, db)
    rows = _scope_rows(scope, db)
    if format == "yaml":
        content = yamllib.safe_dump({scope: rows}, allow_unicode=True, sort_keys=False)
        media, ext = "application/x-yaml", "yaml"
    else:
        content = jsonlib.dumps({scope: rows}, ensure_ascii=False, indent=2)
        media, ext = "application/json", "json"
    return StreamingResponse(io.BytesIO(content.encode("utf-8")), media_type=media,
                             headers={"Content-Disposition": f'attachment; filename="{scope}.{ext}"'})


# ── Personal：profile / notifications ────────────────────────

DEFAULT_NOTIFICATIONS = {
    "job_failure": True, "job_success": False,
    "batch_upload_result": True, "weekly_digest": False,
    "channel_email": True, "channel_slack": False,
}


@router.get("/personal")
def get_personal(db: Session = Depends(get_db)):
    row = db.get(UserPref, "default")
    return {"profile": (row.profile if row else {}) or {},
            "notifications": {**DEFAULT_NOTIFICATIONS, **((row.notifications if row else {}) or {})}}


@router.put("/personal")
def put_personal(body: dict[str, Any], db: Session = Depends(get_db)):
    row = db.get(UserPref, "default")
    if not row:
        row = UserPref(user_id="default", profile={}, notifications={})
        db.add(row)
    if "profile" in body:
        row.profile = {**(row.profile or {}), **(body["profile"] or {})}
    if "notifications" in body:
        row.notifications = {**(row.notifications or {}), **(body["notifications"] or {})}
    db.commit()
    return get_personal(db)
